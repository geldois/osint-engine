from __future__ import annotations

import csv
import io
import re
import zipfile

import pytest
from openpyxl import Workbook

from osint_engine.application.errors.spreadsheet_ingestion_error import (
    FieldTooLargeError,
    FileTooLargeError,
    MalformedSpreadsheetError,
    TooManyRowsError,
    UnsupportedFileTypeError,
)
from osint_engine.domain.value_objects.text_pattern import TextPatternName
from osint_engine.infrastructure.text_ingestion import spreadsheet_reader
from osint_engine.infrastructure.text_ingestion.spreadsheet_reader import (
    read_spreadsheet_text,
)

_VALID_CPF = "11144477735"


def _xlsx_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _inject_cached_formula(
    content: bytes, *, cell_ref: str, formula: str, cached_value: str
) -> bytes:
    with zipfile.ZipFile(io.BytesIO(content)) as source:
        names = source.namelist()
        sheet_xml = source.read("xl/worksheets/sheet1.xml").decode("utf-8")

    pattern = re.compile(rf'(<c r="{cell_ref}"[^>]*>)<v>[^<]*</v>(</c>)')
    replaced = pattern.sub(
        rf"\1<f>{formula}</f><v>{cached_value}</v>\2", sheet_xml, count=1
    )
    assert replaced != sheet_xml

    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(output, "w") as target,
    ):
        for name in names:
            payload = (
                replaced.encode("utf-8")
                if name == "xl/worksheets/sheet1.xml"
                else source.read(name)
            )
            target.writestr(name, payload)

    return output.getvalue()


def _drop_zip_member(content: bytes, *, member: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(content)) as source:
        names = [name for name in source.namelist() if name != member]
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as target:
            for name in names:
                target.writestr(name, source.read(name))

    return output.getvalue()


def _truncate_zip_member(content: bytes, *, member: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(content)) as source:
        names = source.namelist()
        payload = source.read(member)
        truncated = payload[: len(payload) // 2]
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as target:
            for name in names:
                target.writestr(
                    name, truncated if name == member else source.read(name)
                )

    return output.getvalue()


class TestReadSpreadsheetTextCsv:
    def test_three_lines_are_joined_by_newline_cells_by_space(self) -> None:
        content = b"a,b,c\nd,e,f\ng,h,i"

        text = read_spreadsheet_text(content=content, filename="data.csv")

        assert text == "a b c\nd e f\ng h i"

    def test_latin1_bytes_decode_via_fallback(self) -> None:
        content = "café,ação".encode("latin-1")

        text = read_spreadsheet_text(content=content, filename="data.csv")

        assert text == "café ação"

    def test_empty_field_between_two_values_still_produces_a_separator(self) -> None:
        content = b"a,,c"

        text = read_spreadsheet_text(content=content, filename="data.csv")

        assert text == "a  c"

    def test_field_above_the_csv_module_limit_raises_field_too_large_error(
        self,
    ) -> None:
        content = b"a," + b"x" * 200_000

        with pytest.raises(FieldTooLargeError) as exc_info:
            read_spreadsheet_text(content=content, filename="data.csv")

        assert exc_info.value.max_field_bytes == csv.field_size_limit()

    def test_too_many_rows_raises_citing_no_sheet_name(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setattr(spreadsheet_reader, "_MAX_ROWS_PER_SHEET", 2)
        content = b"a\nb\nc"

        with pytest.raises(TooManyRowsError) as exc_info:
            read_spreadsheet_text(content=content, filename="data.csv")

        assert exc_info.value.sheet_name is None


class TestReadSpreadsheetTextXlsx:
    def test_all_sheets_are_scanned_not_only_the_active_one(self) -> None:
        workbook = Workbook()
        first = workbook.active
        assert first is not None
        first.title = "First"
        first["A1"] = "alpha"
        second = workbook.create_sheet("Second")
        second["A1"] = "beta"

        text = read_spreadsheet_text(
            content=_xlsx_bytes(workbook), filename="data.xlsx"
        )

        assert "alpha" in text
        assert "beta" in text

    def test_integer_cell_and_float_cell_both_yield_the_full_cpf(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = int(_VALID_CPF)
        sheet["A2"] = float(_VALID_CPF)

        text = read_spreadsheet_text(
            content=_xlsx_bytes(workbook), filename="data.xlsx"
        )

        matches = TextPatternName.CPF_LOOSE.value.regex.findall(text)

        assert matches.count(_VALID_CPF) == 2

    def test_formula_cell_is_read_by_its_cached_value_not_the_formula_string(
        self,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = 42
        content = _inject_cached_formula(
            _xlsx_bytes(workbook),
            cell_ref="A1",
            formula="CONCATENATE(1,2)",
            cached_value="42",
        )

        text = read_spreadsheet_text(content=content, filename="data.xlsx")

        assert text == "42"
        assert "CONCATENATE" not in text

    def test_none_cell_is_omitted_not_rendered_as_the_string_none(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = "a"
        sheet["D1"] = "d"

        text = read_spreadsheet_text(
            content=_xlsx_bytes(workbook), filename="data.xlsx"
        )

        assert text == "a d"
        assert "None" not in text

    def test_too_many_rows_raises_citing_the_sheet_name(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setattr(spreadsheet_reader, "_MAX_ROWS_PER_SHEET", 2)
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Planilha1"
        for value in ("a", "b", "c"):
            sheet.append([value])

        with pytest.raises(TooManyRowsError) as exc_info:
            read_spreadsheet_text(content=_xlsx_bytes(workbook), filename="data.xlsx")

        assert exc_info.value.sheet_name == "Planilha1"

    def test_content_that_is_not_a_zip_at_all_raises_malformed_spreadsheet_error(
        self,
    ) -> None:
        with pytest.raises(MalformedSpreadsheetError):
            read_spreadsheet_text(content=b"not a zip file", filename="data.xlsx")

    def test_zip_missing_content_types_raises_malformed_spreadsheet_error(
        self,
    ) -> None:
        workbook = Workbook()
        content = _drop_zip_member(_xlsx_bytes(workbook), member="[Content_Types].xml")

        with pytest.raises(MalformedSpreadsheetError):
            read_spreadsheet_text(content=content, filename="data.xlsx")

    def test_zip_missing_workbook_part_raises_malformed_spreadsheet_error(
        self,
    ) -> None:
        workbook = Workbook()
        content = _drop_zip_member(_xlsx_bytes(workbook), member="xl/workbook.xml")

        with pytest.raises(MalformedSpreadsheetError):
            read_spreadsheet_text(content=content, filename="data.xlsx")

    def test_truncated_sheet_xml_raises_malformed_spreadsheet_error(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet["A1"] = "alpha"
        content = _truncate_zip_member(
            _xlsx_bytes(workbook), member="xl/worksheets/sheet1.xml"
        )

        with pytest.raises(MalformedSpreadsheetError):
            read_spreadsheet_text(content=content, filename="data.xlsx")


class TestReadSpreadsheetTextLimits:
    def test_file_above_10mb_raises_before_parsing(self) -> None:
        content = b"x" * (10 * 1024 * 1024 + 1)

        with pytest.raises(FileTooLargeError) as exc_info:
            read_spreadsheet_text(content=content, filename="data.xlsx")

        assert exc_info.value.actual_bytes == len(content)

    @pytest.mark.parametrize("filename", ["data.xls", "data.xlsm", "data.ods", "data"])
    def test_unsupported_extension_raises_without_opening_content(
        self, filename: str
    ) -> None:
        with pytest.raises(UnsupportedFileTypeError):
            read_spreadsheet_text(content=b"irrelevant", filename=filename)
